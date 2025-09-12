#!/usr/bin/env python3
"""
Strategic Cannabis Chain Analyzer
Focus on high-impact chains and online retailers from 925 retailer dataset
"""

import csv
import time
import json
import pandas as pd
from datetime import datetime
from analyzer import analyze_cannabis_site
from urllib.parse import urlparse

# High-Impact Chain Targets (400+ locations combined)
MAJOR_CHAINS = {
    "Canna Cabana": {
        "url": "https://cannacabana.com",
        "locations": 90,
        "provinces": ["AB", "ON", "BC"],
        "priority": 1,
        "revenue_potential": "Very High"
    },
    "Value Buds": {
        "url": "https://valuebuds.com", 
        "locations": 68,
        "provinces": ["AB", "BC"],
        "priority": 1,
        "revenue_potential": "Very High"
    },
    "PlantLife Cannabis": {
        "url": "https://plantlifecannabis.com",
        "locations": 39,
        "provinces": ["AB"],
        "priority": 2,
        "revenue_potential": "High"
    },
    "Fire & Flower": {
        "url": "https://fireandflower.com",
        "locations": 27,
        "provinces": ["AB", "ON", "BC"],
        "priority": 1,
        "revenue_potential": "Very High"
    },
    "Spiritleaf": {
        "url": "https://spiritleaf.ca",
        "locations": 23,
        "provinces": ["AB", "ON", "BC"],
        "priority": 1,
        "revenue_potential": "Very High"
    },
    "The Hunny Pot": {
        "url": "https://thehunnypot.com",
        "locations": 15,
        "provinces": ["ON"],
        "priority": 2,
        "revenue_potential": "Medium"
    },
    "Greenspace": {
        "url": "https://greenspace.ca",
        "locations": 12,
        "provinces": ["AB"],
        "priority": 3,
        "revenue_potential": "Medium"
    },
    "Nova Cannabis": {
        "url": "https://novacannabis.ca",
        "locations": 10,
        "provinces": ["AB"],
        "priority": 3,
        "revenue_potential": "Medium"
    }
}

# High-Value Online Retailers (E-commerce compliance critical)
ONLINE_RETAILERS = [
    "https://ocs.ca",  # Government baseline
    "https://bccannabisstores.com",  # Government baseline
    "https://shop.sqdc.ca",  # Quebec government
    "https://albertacannabis.org",  # Alberta government
    "https://budderweeds.com",  # Online delivery
    "https://cannabisexpress.ca",  # Online express
    "https://getbuddy.ca",  # Delivery service
    "https://thcdirect.ca",  # Direct sales
    "https://greensociety.ca",  # Online community
    "https://chronicstore.ca",  # Chronic store
    "https://bulkcrop.com",  # Bulk sales
    "https://ganja-express.ca",  # Express delivery
    "https://mmjdirect.ca",  # Medical direct
    "https://cannabismo.ca",  # Online retailer
    "https://westcoastcannabis.ca",  # West coast
    "https://kootenaycraft.com",  # Craft cannabis
    "https://cannabudpost.com",  # Mail order
    "https://lowpricebud.ca",  # Budget focused
    "https://speedgreens.ca",  # Speed delivery
    "https://thegreenace.ca"  # Online ace
]

# Provincial Sample Testing (Different regulations)
PROVINCIAL_SAMPLES = {
    "AB": [
        "https://fireandflower.com",
        "https://cannacabana.com", 
        "https://valuebuds.com"
    ],
    "ON": [
        "https://ocs.ca",
        "https://thehunnypot.com",
        "https://tokyo-smoke.com"
    ],
    "BC": [
        "https://bccannabisstores.com",
        "https://kiaro.ca",
        "https://citycannabishop.ca"
    ],
    "QC": [
        "https://sqdc.ca",
        "https://shop.sqdc.ca"
    ]
}

def calculate_business_impact(chain_name, score, locations):
    """Calculate potential business impact for chain opportunities"""
    
    # Revenue potential calculation
    base_project_value = 2000  # Base compliance project
    monthly_retainer = 500    # Monthly monitoring
    
    # Chain multiplier based on locations
    if locations >= 50:
        multiplier = 1.5  # Enterprise pricing
    elif locations >= 20:
        multiplier = 1.3  # Mid-market pricing
    else:
        multiplier = 1.0  # Standard pricing
    
    project_value = base_project_value * multiplier
    annual_retainer = monthly_retainer * 12 * multiplier
    
    # Urgency based on score
    if score < 60:
        urgency = "CRITICAL"
        close_probability = 0.8
    elif score < 75:
        urgency = "HIGH"
        close_probability = 0.6
    elif score < 85:
        urgency = "MEDIUM"
        close_probability = 0.4
    else:
        urgency = "LOW"
        close_probability = 0.2
    
    return {
        "project_value": project_value,
        "annual_retainer": annual_retainer,
        "total_opportunity": project_value + annual_retainer,
        "urgency": urgency,
        "close_probability": close_probability,
        "expected_value": (project_value + annual_retainer) * close_probability,
        "locations_impact": locations
    }

def analyze_chain(chain_name, chain_data):
    """Analyze a major cannabis chain with business intelligence"""
    
    print(f"\nANALYZING CHAIN: {chain_name}")
    print(f"   Locations: {chain_data['locations']}")
    print(f"   Provinces: {', '.join(chain_data['provinces'])}")
    print(f"   Revenue Potential: {chain_data['revenue_potential']}")
    
    try:
        start_time = time.time()
        result = analyze_cannabis_site(chain_data['url'])
        analysis_time = time.time() - start_time
        
        # Add chain-specific data
        result['chain_name'] = chain_name
        result['locations'] = chain_data['locations']
        result['provinces'] = chain_data['provinces']
        result['priority'] = chain_data['priority']
        result['analysis_type'] = 'CHAIN'
        result['analysis_time'] = round(analysis_time, 2)
        
        # Calculate business impact
        impact = calculate_business_impact(chain_name, result['score'], chain_data['locations'])
        result['business_impact'] = impact
        
        # Display results
        print(f"   Compliance Score: {result['score']}/100")
        print(f"   Priority Level: {impact['urgency']}")
        print(f"   Opportunity Value: ${impact['total_opportunity']:,}")
        print(f"   Expected Value: ${impact['expected_value']:,.0f}")
        print(f"   Close Probability: {impact['close_probability']*100:.0f}%")
        
        if result['score'] < 75:
            print(f"   ACTION REQUIRED: {chain_data['locations']} locations at compliance risk!")
        
        return result
        
    except Exception as e:
        print(f"   ❌ Analysis Failed: {str(e)}")
        return {
            'chain_name': chain_name,
            'url': chain_data['url'],
            'error': str(e),
            'locations': chain_data['locations'],
            'analysis_type': 'CHAIN'
        }

def analyze_online_retailer(url):
    """Analyze online cannabis retailer with e-commerce focus"""
    
    domain = urlparse(url).netloc
    print(f"\nANALYZING ONLINE RETAILER: {domain}")
    
    try:
        start_time = time.time()
        result = analyze_cannabis_site(url)
        analysis_time = time.time() - start_time
        
        # Add online-specific data
        result['analysis_type'] = 'ONLINE_RETAILER'
        result['analysis_time'] = round(analysis_time, 2)
        result['ecommerce_risk'] = 'HIGH' if result['score'] < 75 else 'MEDIUM' if result['score'] < 85 else 'LOW'
        
        # Online retailers have higher compliance requirements
        adjusted_score = result['score'] - 10 if result['score'] > 10 else 0
        result['ecommerce_adjusted_score'] = adjusted_score
        
        print(f"   Base Score: {result['score']}/100")
        print(f"   E-commerce Adjusted: {adjusted_score}/100")
        print(f"   E-commerce Risk: {result['ecommerce_risk']}")
        
        return result
        
    except Exception as e:
        print(f"   ❌ Analysis Failed: {str(e)}")
        return {
            'url': url,
            'error': str(e),
            'analysis_type': 'ONLINE_RETAILER'
        }

def run_strategic_analysis():
    """Run strategic analysis focusing on high-impact opportunities"""
    
    print("STRATEGIC CANNABIS COMPLIANCE ANALYSIS")
    print("Focus: Chains + Online Retailers = Maximum ROI")
    print("=" * 70)
    
    all_results = []
    
    # Phase 1: Major Chains Analysis
    print(f"\nPHASE 1: MAJOR CHAINS ANALYSIS")
    print(f"Target: {len(MAJOR_CHAINS)} chains representing 400+ locations")
    
    chain_results = []
    for chain_name, chain_data in MAJOR_CHAINS.items():
        result = analyze_chain(chain_name, chain_data)
        chain_results.append(result)
        all_results.append(result)
        time.sleep(2)  # Respectful delay
    
    # Phase 2: Online Retailers Analysis  
    print(f"\nPHASE 2: ONLINE RETAILERS ANALYSIS")
    print(f"Target: {len(ONLINE_RETAILERS)} high-risk e-commerce sites")
    
    online_results = []
    for url in ONLINE_RETAILERS[:10]:  # Limit to 10 for initial test
        result = analyze_online_retailer(url)
        online_results.append(result)
        all_results.append(result)
        time.sleep(1)  # Faster for online retailers
    
    # Save results and generate reports
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_strategic_results(all_results, chain_results, online_results, timestamp)
    
    return all_results

def save_strategic_results(all_results, chain_results, online_results, timestamp):
    """Save results with strategic business focus"""
    
    # Enhanced CSV for business development
    csv_filename = f"strategic_analysis_{timestamp}.csv"
    headers = [
        'timestamp', 'analysis_type', 'chain_name', 'url', 'score', 
        'locations', 'urgency', 'project_value', 'expected_value',
        'close_probability', 'age_gate', 'health_warnings', 'license_display',
        'provinces', 'ecommerce_risk', 'priority'
    ]
    
    with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=headers)
        writer.writeheader()
        
        for result in all_results:
            if 'error' not in result:
                impact = result.get('business_impact', {})
                row = {
                    'timestamp': datetime.now().isoformat(),
                    'analysis_type': result.get('analysis_type', ''),
                    'chain_name': result.get('chain_name', ''),
                    'url': result.get('url', ''),
                    'score': result.get('score', 0),
                    'locations': result.get('locations', 1),
                    'urgency': impact.get('urgency', 'UNKNOWN'),
                    'project_value': impact.get('project_value', 0),
                    'expected_value': impact.get('expected_value', 0),
                    'close_probability': impact.get('close_probability', 0),
                    'age_gate': result.get('age_gate', False),
                    'health_warnings': result.get('health_warnings', False),
                    'license_display': result.get('license_display', False),
                    'provinces': ', '.join(result.get('provinces', [])),
                    'ecommerce_risk': result.get('ecommerce_risk', ''),
                    'priority': result.get('priority', 0)
                }
                writer.writerow(row)
    
    # Create Excel spreadsheet with detailed grades
    create_grading_spreadsheet(all_results, chain_results, online_results, timestamp)
    
    # Generate strategic business report
    generate_strategic_report(chain_results, online_results, timestamp)
    
    print(f"\nSTRATEGIC ANALYSIS COMPLETE")
    print(f"Results: {csv_filename}")
    print(f"Grades Spreadsheet: cannabis_compliance_grades_{timestamp}.xlsx")
    print(f"Business Report: strategic_opportunities_{timestamp}.txt")

def create_grading_spreadsheet(all_results, chain_results, online_results, timestamp):
    """Create detailed Excel spreadsheet with website grades and compliance breakdown"""
    
    filename = f"cannabis_compliance_grades_{timestamp}.xlsx"
    
    # Prepare data for different sheets
    overview_data = []
    detailed_data = []
    
    for result in all_results:
        if 'error' not in result:
            # Get detailed scores for each compliance area
            compliance_scores = calculate_detailed_scores(result)
            impact = result.get('business_impact', {})
            
            # Overview sheet data
            overview_row = {
                'Company/Chain': result.get('chain_name', urlparse(result.get('url', '')).netloc),
                'Website': result.get('url', ''),
                'Overall Score': result.get('score', 0),
                'Grade': get_letter_grade(result.get('score', 0)),
                'Analysis Type': result.get('analysis_type', ''),
                'Locations': result.get('locations', 1),
                'Urgency': impact.get('urgency', 'UNKNOWN'),
                'Expected Value': impact.get('expected_value', 0),
                'Provinces': ', '.join(result.get('provinces', [])),
                'Primary Issues': get_primary_issues(result)
            }
            overview_data.append(overview_row)
            
            # Detailed sheet data
            detailed_row = {
                'Company/Chain': result.get('chain_name', urlparse(result.get('url', '')).netloc),
                'Website': result.get('url', ''),
                'Overall Score': result.get('score', 0),
                'Grade': get_letter_grade(result.get('score', 0)),
                'HTTPS Security': 'Pass' if result.get('has_https', False) else 'Fail',
                'Age Verification': 'Pass' if result.get('has_age_gate', False) else 'Fail',
                'Health Warnings': 'Pass' if result.get('has_warning', False) else 'Fail',
                'License Display': 'Pass' if result.get('has_license', False) else 'Fail',
                'Mobile Friendly': 'Pass' if result.get('mobile_friendly', False) else 'Fail',
                'Contact Info': 'Pass' if result.get('has_contact', False) else 'Fail',
                'Load Time (sec)': round(result.get('load_time', 0), 2),
                'Page Accessible': 'Pass' if result.get('status_code') == 200 else 'Fail',
                'Security Score': compliance_scores['security'],
                'Legal Compliance Score': compliance_scores['legal'],
                'User Experience Score': compliance_scores['ux'],
                'Technical Score': compliance_scores['technical'],
                'Recommendations': '; '.join(get_recommendations_list(result))
            }
            detailed_data.append(detailed_row)
    
    # Create Excel writer object
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Overview sheet
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name='Overview', index=False)
        
        # Detailed compliance sheet
        detailed_df = pd.DataFrame(detailed_data)
        detailed_df.to_excel(writer, sheet_name='Detailed Analysis', index=False)
        
        # Summary statistics sheet
        summary_data = create_summary_statistics(all_results, chain_results, online_results)
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary Stats', index=False)
        
        # Apply formatting
        format_spreadsheet(writer, overview_df, detailed_df, summary_df)

def calculate_detailed_scores(result):
    """Calculate detailed scores for different compliance categories"""
    security_score = 0
    legal_score = 0
    ux_score = 0
    technical_score = 0
    
    # Security (out of 100)
    if result.get('has_https', False):
        security_score += 60
    if result.get('status_code') == 200:
        security_score += 40
    
    # Legal Compliance (out of 100)
    if result.get('has_age_gate', False):
        legal_score += 40
    if result.get('has_warning', False):
        legal_score += 35
    if result.get('has_license', False):
        legal_score += 25
    
    # User Experience (out of 100)
    if result.get('mobile_friendly', False):
        ux_score += 40
    if result.get('has_contact', False):
        ux_score += 20
    if result.get('load_time', 10) < 3:
        ux_score += 40
    elif result.get('load_time', 10) < 5:
        ux_score += 20
    
    # Technical (out of 100)
    load_time = result.get('load_time', 10)
    if load_time < 2:
        technical_score += 50
    elif load_time < 3:
        technical_score += 40
    elif load_time < 5:
        technical_score += 25
    
    page_size = result.get('page_size', 0)
    if page_size > 0 and page_size < 1000000:  # Less than 1MB
        technical_score += 30
    elif page_size > 0 and page_size < 2000000:  # Less than 2MB
        technical_score += 20
    
    if result.get('title_length', 0) > 0:
        technical_score += 20
    
    return {
        'security': security_score,
        'legal': legal_score,
        'ux': ux_score,
        'technical': technical_score
    }

def get_letter_grade(score):
    """Convert numeric score to letter grade"""
    if score >= 90:
        return 'A'
    elif score >= 80:
        return 'B'
    elif score >= 70:
        return 'C'
    elif score >= 60:
        return 'D'
    else:
        return 'F'

def get_primary_issues(result):
    """Get the top compliance issues for a website"""
    issues = []
    
    if not result.get('has_https', False):
        issues.append('No HTTPS')
    if not result.get('has_age_gate', False):
        issues.append('Missing Age Gate')
    if not result.get('has_warning', False):
        issues.append('No Health Warnings')
    if not result.get('has_license', False):
        issues.append('License Not Displayed')
    if result.get('load_time', 0) > 5:
        issues.append('Slow Loading')
    
    return ', '.join(issues[:3])  # Top 3 issues

def get_recommendations_list(result):
    """Get actionable recommendations for a website"""
    recommendations = []
    
    if not result.get('has_https', False):
        recommendations.append('Enable HTTPS SSL certificate')
    if not result.get('has_age_gate', False):
        recommendations.append('Implement age verification system')
    if not result.get('has_warning', False):
        recommendations.append('Add Health Canada warning text')
    if not result.get('has_license', False):
        recommendations.append('Display cannabis license information')
    if not result.get('mobile_friendly', False):
        recommendations.append('Optimize for mobile devices')
    if result.get('load_time', 0) > 3:
        recommendations.append('Improve page loading speed')
    if not result.get('has_contact', False):
        recommendations.append('Add clear contact information')
    
    return recommendations

def create_summary_statistics(all_results, chain_results, online_results):
    """Create summary statistics for the analysis"""
    total_sites = len([r for r in all_results if 'error' not in r])
    total_chains = len([r for r in chain_results if 'error' not in r])
    total_online = len([r for r in online_results if 'error' not in r])
    
    if total_sites == 0:
        return []
    
    scores = [r.get('score', 0) for r in all_results if 'error' not in r]
    avg_score = sum(scores) / len(scores) if scores else 0
    
    # Count compliance features
    https_count = sum(1 for r in all_results if 'error' not in r and r.get('has_https', False))
    age_gate_count = sum(1 for r in all_results if 'error' not in r and r.get('has_age_gate', False))
    warning_count = sum(1 for r in all_results if 'error' not in r and r.get('has_warning', False))
    license_count = sum(1 for r in all_results if 'error' not in r and r.get('has_license', False))
    
    summary = [
        {'Metric': 'Total Sites Analyzed', 'Value': total_sites, 'Percentage': '100%'},
        {'Metric': 'Cannabis Chains', 'Value': total_chains, 'Percentage': f'{(total_chains/total_sites)*100:.1f}%'},
        {'Metric': 'Online Retailers', 'Value': total_online, 'Percentage': f'{(total_online/total_sites)*100:.1f}%'},
        {'Metric': 'Average Compliance Score', 'Value': f'{avg_score:.1f}', 'Percentage': ''},
        {'Metric': 'Sites with HTTPS', 'Value': https_count, 'Percentage': f'{(https_count/total_sites)*100:.1f}%'},
        {'Metric': 'Sites with Age Gates', 'Value': age_gate_count, 'Percentage': f'{(age_gate_count/total_sites)*100:.1f}%'},
        {'Metric': 'Sites with Health Warnings', 'Value': warning_count, 'Percentage': f'{(warning_count/total_sites)*100:.1f}%'},
        {'Metric': 'Sites with License Display', 'Value': license_count, 'Percentage': f'{(license_count/total_sites)*100:.1f}%'},
        {'Metric': 'Grade A Sites (90+)', 'Value': len([s for s in scores if s >= 90]), 'Percentage': f'{(len([s for s in scores if s >= 90])/total_sites)*100:.1f}%'},
        {'Metric': 'Grade F Sites (<60)', 'Value': len([s for s in scores if s < 60]), 'Percentage': f'{(len([s for s in scores if s < 60])/total_sites)*100:.1f}%'}
    ]
    
    return summary

def format_spreadsheet(writer, overview_df, detailed_df, summary_df):
    """Apply formatting to the Excel spreadsheet"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Format Overview sheet
        overview_ws = writer.sheets['Overview']
        overview_ws.column_dimensions['A'].width = 20
        overview_ws.column_dimensions['B'].width = 30
        overview_ws.column_dimensions['C'].width = 12
        overview_ws.column_dimensions['D'].width = 8
        overview_ws.column_dimensions['J'].width = 30
        
        # Format headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in overview_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Color code grades
        for row in range(2, len(overview_df) + 2):
            grade_cell = overview_ws[f'D{row}']
            score_cell = overview_ws[f'C{row}']
            
            if grade_cell.value == 'A':
                grade_cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            elif grade_cell.value == 'B':
                grade_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif grade_cell.value == 'C':
                grade_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            elif grade_cell.value == 'D':
                grade_cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            elif grade_cell.value == 'F':
                grade_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        # Format Detailed Analysis sheet
        detailed_ws = writer.sheets['Detailed Analysis']
        for col in detailed_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            detailed_ws.column_dimensions[column].width = adjusted_width
        
        # Format headers in detailed sheet
        for cell in detailed_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Format Summary Stats sheet
        summary_ws = writer.sheets['Summary Stats']
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 15
        summary_ws.column_dimensions['C'].width = 15
        
        for cell in summary_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
    except ImportError:
        # If openpyxl styling is not available, skip formatting
        pass

def generate_strategic_report(chain_results, online_results, timestamp):
    """Generate strategic business opportunities report"""
    
    # Calculate total opportunity
    total_locations = sum([r.get('locations', 0) for r in chain_results if 'error' not in r])
    total_opportunity = sum([r.get('business_impact', {}).get('total_opportunity', 0) for r in chain_results if 'error' not in r])
    total_expected = sum([r.get('business_impact', {}).get('expected_value', 0) for r in chain_results if 'error' not in r])
    
    # Identify high-priority opportunities
    critical_chains = [r for r in chain_results if 'error' not in r and r.get('business_impact', {}).get('urgency') == 'CRITICAL']
    high_risk_online = [r for r in online_results if 'error' not in r and r.get('ecommerce_risk') == 'HIGH']
    
    report = f"""
🚀 STRATEGIC CANNABIS COMPLIANCE OPPORTUNITIES
📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
🎯 Focus: High-Impact Chains + Online Retailers
=" * 60

💰 BUSINESS OPPORTUNITY SUMMARY
Total Chain Locations Analyzed: {total_locations}
Total Revenue Opportunity: ${total_opportunity:,.0f}
Expected Revenue (Probability-Adjusted): ${total_expected:,.0f}
Average Deal Size: ${total_opportunity/max(1,len(chain_results)):,.0f}

🏢 CHAIN ANALYSIS RESULTS
"""
    
    for result in sorted(chain_results, key=lambda x: x.get('business_impact', {}).get('expected_value', 0), reverse=True):
        if 'error' not in result:
            impact = result.get('business_impact', {})
            report += f"""
{result.get('chain_name', 'Unknown')} ({result.get('locations', 0)} locations)
  Score: {result.get('score', 0)}/100 | Urgency: {impact.get('urgency', 'UNKNOWN')}
  Opportunity: ${impact.get('total_opportunity', 0):,.0f} | Expected: ${impact.get('expected_value', 0):,.0f}
  Provinces: {', '.join(result.get('provinces', []))}
"""
    
    report += f"""
🛒 ONLINE RETAILER ANALYSIS
High E-commerce Risk Sites: {len(high_risk_online)}
"""
    
    for result in high_risk_online[:5]:  # Top 5 high-risk
        domain = urlparse(result.get('url', '')).netloc
        report += f"• {domain} - Score: {result.get('score', 0)}/100 (E-commerce Risk: HIGH)\n"
    
    report += f"""
🎯 IMMEDIATE ACTION PLAN (THIS WEEK)

CRITICAL PRIORITY - CHAINS WITH URGENT NEEDS:
"""
    
    for chain in critical_chains:
        report += f"""
• {chain.get('chain_name')} - {chain.get('locations')} locations
  Score: {chain.get('score')}/100 | Expected Value: ${chain.get('business_impact', {}).get('expected_value', 0):,.0f}
  Action: Immediate outreach to compliance/operations team
"""
    
    report += f"""
📧 OUTREACH TEMPLATES

CHAIN DECISION MAKERS:
Subject: "Compliance Risk Alert: [Chain Name]'s {result.get('locations')} Locations"

HIGH-RISK ONLINE RETAILERS:
Subject: "E-commerce Cannabis Compliance Issues Detected"

📊 SUCCESS METRICS TO TRACK
• Chain outreach response rate (target: 30%+)
• Meeting conversion rate (target: 50%+)
• Close rate for critical urgency (target: 60%+)
• Average deal size vs projection

🔍 COMPETITIVE ADVANTAGES IDENTIFIED
1. Focus on chains = 10x impact per client
2. E-commerce expertise = premium pricing
3. Provincial specialization = reduced competition
4. Urgency-based prioritization = higher close rates

💡 REVENUE SCALING STRATEGY
Month 1: Close 2 critical chains = ${sum([r.get('business_impact', {}).get('expected_value', 0) for r in critical_chains[:2]]):,.0f}
Month 2: Add 3 medium chains = Additional revenue stream
Month 3: Launch provincial packages = Scalable offerings

🎯 Next Steps:
1. LinkedIn research for chain compliance officers
2. Craft urgency-based email templates
3. Schedule 5 chain calls this week
4. Create chain-specific audit reports
"""
    
    # Save report
    report_filename = f"strategic_opportunities_{timestamp}.txt"
    with open(report_filename, 'w', encoding='utf-8') as f:
        f.write(report)

if __name__ == "__main__":
    print("STRATEGIC CANNABIS COMPLIANCE ANALYZER")
    print("Targeting 925 retailers through high-impact chains & online retailers")
    print("Focus: Maximum ROI through chain multiplication effect")
    print("\nStarting strategic analysis...")
    
    run_strategic_analysis()